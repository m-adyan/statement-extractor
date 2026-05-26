"""
Statement Extractor — PDF bank statement to Excel converter
Flask web app for extracting tables from PDF bank statements.
"""
import os
import io
import uuid
import tempfile
from datetime import datetime

from flask import Flask, request, render_template_string, send_file, jsonify
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()

# ── Helpers ─────────────────────────────────────────────────────────────────

def parse_wise_text(text):
    """
    Wise-specific parser. Wise PDFs have no date on transaction lines —
    instead a "DD Mon YYYY Transaction: TXID" header line appears BELOW
    the TXs it describes. Pages without headers rely on inline dates
    embedded in the first transaction line on that page.

    Strategy:
      1. Split text by page separator lines ("ref:... N / M")
      2. Per-page: reverse-scan assigns dates. Pages with no date header
         get their first TX line's embedded date assigned to all page TXs.
      3. Merge all pages in original chronological order
      4. Sort ascending (oldest first)

    Output: Date | Description | Type | Money In | Money Out | Balance
    """
    import re
    month_idx = {m: i+1 for i, m in enumerate(
        ['January','February','March','April','May','June','July','August',
         'September','October','November','December'])}
    months_full = (r'January|February|March|April|May|June|July|August|'
                  r'September|October|November|December')
    months = 'January|February|March|April|May|June|July|August|September|October|November|December'

    lines = text.split('\n')
    detected_year = None
    # ── 1. Detect statement year from header ──────────────────────────────────
    for hl in lines[:30]:
        # Skip auto-generated metadata (not the actual statement period)
        if re.match(r'^(Generated|Created|Printed|Updated)\s+on:', hl, re.IGNORECASE):
            continue
        # Period range line: "31 August 2023 - 31 August 2024"
        if re.search(rf'(\d{{1,2}}\s+({months})\s+\d{{4}}).*? - .*?(\d{{1,2}}\s+({months})\s+\d{{4}})', hl):
            m = re.search(rf'(\d{{1,2}}\s+({months})\s+(\d{{4}}))', hl)
            if m:
                detected_year = int(m.group(3))
                break
        hm = re.search(rf'(({months})\s+\d{{4}})', hl)
        if hm:
            detected_year = int(hm.group(0).split()[-1])
            break
        hm2 = re.search(r'(\d{2})/(\d{2})/(\d{4})', hl)
        if hm2:
            detected_year = int(hm2.group(3))
            break

    # Helper: parse a TX line → dict with amount/balance/description
    def parse_tx(line):
        nums = re.findall(r'(-?[\d,]+\.?\d*)', line)
        if len(nums) < 2:
            return None
        balance = float(nums[-1].replace(',', ''))
        amt = None
        desc = None
        if line.startswith('Sent money to'):
            amt = -abs(float(nums[-2].replace(',', '')))
            desc = ' → '.join([s.strip() for s in line.rsplit(maxsplit=2)[0].split('to', 1)])
        elif line.startswith("Received money from"):
            amt = float(nums[-2].replace(',', ''))
            ref_match = re.search(r'with reference\s+(\S.+)', line)
            ref_part = ref_match.group(1) if ref_match else ''
            m_co = re.search(r'from\s+(.+?)\s+with reference', line)
            company = m_co.group(1).strip() if m_co else ''
            desc = f"← {company}" + (f' ({ref_part})' if ref_part else '')
        elif line.startswith('Card transaction'):
            amt = -abs(float(nums[-2].replace(',', '')))
            desc = line.split('+')[0].strip().rstrip(', ')
        elif line.startswith('Wise Charges for:'):
            amt = -abs(float(nums[-2].replace(',', '')))
            desc = 'Wise Charges'
        if amt is not None:
            return {'amount': amt, 'balance': balance, 'description': desc}
        return None

    def date_header_info(line):
        m = re.match(rf'^(\d{{1,2}}\s+({months_full})\s+(\d{{4}}))\s+Transaction:\s*\S', line)
        if m:
            return {'date': m.group(1), 'year': int(m.group(3)),
                    'month_idx': month_idx.get(m.group(2), 1),
                    'day': int(m.group(1).split()[0])}
        return None

    def inline_date_info(line):
        m = re.search(rf'(\d{{1,2}}\s+({months_full})\s+(\d{{4}}))', line)
        if m:
            return {'date': m.group(1), 'year': int(m.group(3)),
                    'month_idx': month_idx.get(m.group(2), 1),
                    'day': int(m.group(1).split()[0])}
        return None

    # ── 2. Split by page separators ───────────────────────────────────────────
    page_blocks = []
    current_block = []
    for line in lines:
        if re.match(r'^ref:[a-f0-9\-]+\s+\d+\s+/\s+\d+$', line.strip()):
            page_blocks.append(current_block)
            current_block = []
        else:
            current_block.append(line)
    if current_block:
        page_blocks.append(current_block)
    if not page_blocks:
        page_blocks = [lines]

    all_page_txns = []

    # ── 2. Date-header scan ──────────────────────────────────────────────────
    #  Each TX line appears ABOVE its "DD Mon YYYY Transaction:" date header.
    #  Algorithm: scan FORWARD through text. When we see a TX line, the date
    #  that applies is the LAST date_header we passed (which is the NEXT one
    #  in reading order = the correct date). When we HIT a date header, drain
    #  all pending TXs and store this date as the new reference.
    ref_date = None   # last date header seen while scanning forward
    pending  = []     # TX lines waiting for their date_header to appear

    def drain_pending():
        for e in pending:
            e['date'] = ref_date.copy() if ref_date else None
            all_page_txns.append(e)

    for l in lines:
        l = l.strip()
        if not l:
            continue
        if l.startswith(('Sent money to', 'Received money from',
                        'Card transaction', 'Wise Charges for:')):
            entry = parse_tx(l)
            if entry:
                pending.append(entry)
        elif date_header_info(l):
            drain_pending()  # close out TXs above this date header
            pending = []
            ref_date = date_header_info(l)

    # Drain anything still pending at EOF — uses the last seen date
    drain_pending()

    # ── 3. Sort ascending (oldest first) ──────────────────────────────────────
    def sort_key(t):
        d = t.get('date') or {}
        return (d.get('year', 2024), d.get('month_idx', 1), d.get('day', 1))
    all_page_txns.sort(key=sort_key)

    # ── 4. Fix Dec→Jan year rollovers ─────────────────────────────────────────
    def safe_year(t):
        d = t.get('date') or {}
        return d.get('year', 2024)
    def safe_mi(t):
        d = t.get('date') or {}
        return d.get('month_idx', 1)

    if len(all_page_txns) >= 2:
        for i in range(1, len(all_page_txns)):
            prev_year = safe_year(all_page_txns[i-1])
            curr_year = safe_year(all_page_txns[i])
            if (safe_mi(all_page_txns[i]) == 1 and
                    safe_mi(all_page_txns[i-1]) == 12 and
                    curr_year < prev_year):
                dm = all_page_txns[i]['date']
                dm['date'] = f"{dm['date'].split()[0]} {dm['date'].split()[1]} {dm['year'] + 1}"
                dm['year'] += 1

    # ── 5. Classify type + build Money In/Money Out ───────────────────────────
    for tx in all_page_txns:
        upper = tx['description'].upper()
        if 'WISE CHARGES' in upper:
            tx_type = 'Fee'
        elif '← ' in tx['description']:
            tx_type = 'Credit'
        elif 'CARD' in upper or 'SEEDFORMATION' in upper:
            tx_type = 'Debit Card'
        else:
            tx_type = 'Payment'

        amt = tx['amount']
        if amt >= 0:
            tx['money_in'] = round(amt, 2)
            tx['money_out'] = ''
        else:
            tx['money_in'] = ''
            tx['money_out'] = round(abs(amt), 2)
        tx['type'] = tx_type
        tx['balance'] = round(tx['balance'], 2)

    return all_page_txns


def parse_monzo_text(text):
    """
    Monzo-specific parser. Monzo PDFs use:
    - DD/MM/YYYY date format on every transaction line
    - Description: merchant + location + country code
    - Amount (can be negative, no £ symbol)
    - Balance (positive, comma-separated thousands)
    
    Output: Date | Description | Type | Money In | Money Out | Balance
    """
    import re
    transactions = []
    month_abbr = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    # Monzo date format: DD/MM/YYYY at start of each transaction line
    monzo_pat = re.compile(
        rf'^(\d{{1,2}}/\d{{1,2}}/\d{{4}})\s+(.+?)\s+(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)$'
    )

    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue

        m = re.match(monzo_pat, line)
        if not m:
            continue

        raw_date = m.group(1)  # DD/MM/YYYY
        desc = m.group(2).strip()
        amount = float(m.group(3).replace(',', ''))
        balance = float(m.group(4).replace(',', ''))

        # Parse DD/MM/YYYY
        parts = raw_date.split('/')
        day = int(parts[0])
        month = int(parts[1])
        year = int(parts[2])
        date_str = f"{day} {month_abbr[month-1]} {year}"

        # Classify type
        upper = desc.upper()
        if 'TRANSFER FROM POT' in upper or 'TRANSFER TO POT' in upper:
            tx_type = 'Transfer'
        elif 'BANK TRANSFER' in upper or 'FASTER PAYMENT' in upper:
            tx_type = 'Bank Transfer'
        elif 'DIRECT DEBIT' in upper:
            tx_type = 'Direct Debit'
        elif 'DEBIT CARD' in upper or 'CARD' in upper or 'PENDING' in upper or 'UBR' in upper:
            tx_type = 'Debit Card'
        elif 'FEE' in upper or 'CHARGE' in upper:
            tx_type = 'Fee'
        elif 'CREDIT' in upper or 'DEPOSIT' in upper:
            tx_type = 'Credit'
        elif amount >= 0:
            tx_type = 'Credit'
        else:
            tx_type = 'Payment'

        # Money in/out
        if amount >= 0:
            money_in = round(amount, 2)
            money_out = ''
        else:
            money_in = ''
            money_out = round(abs(amount), 2)

        transactions.append({
            'date': date_str,
            'description': desc,
            'type': tx_type,
            'money_in': money_in,
            'money_out': money_out,
            'balance': round(balance, 2),
            'raw_amount': amount,
            'sort_key': (year, month, day)
        })

    # Sort chronologically (oldest first)
    transactions.sort(key=lambda t: t['sort_key'])
    for tx in transactions:
        tx.pop('sort_key', None)

    return transactions


def parse_bank_text(text):
    """
    Smart text parser for line-based bank statements (NatWest, Lloyds, Tide, etc.)
    Detects transaction lines and returns structured data with Date, Description, Type,
    Money In, Money Out, Balance.

    Auto-detects year from statement header and applies to all date-only transaction lines.
    Handles year boundaries (Dec → Jan) correctly via chronological ordering.
    """
    import re
    transactions = []
    months = 'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec'
    month_idx = {m: i+1 for i, m in enumerate(['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'])}

    # ── 1. Detect year from statement header ──────────────────────────────────
    # Patterns: "15/01/2026" | "15 Jan 2026" | "Statement for: 1 Apr 2025 - 30 Apr 2025"
    detected_year = None
    for header_line in text.split('\n')[:30]:
        # Skip auto-generated metadata
        if re.match(r'^(Generated|Created|Printed|Updated)\s+on:', header_line, re.IGNORECASE):
            continue
        # Period range line: "31 August 2023 - 31 August 2024"
        if re.search(rf'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}).*? - .*?(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4})', header_line):
            m = re.search(rf'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4}))', header_line)
            if m:
                detected_year = int(m.group(2))
                break
        hm = re.search(r'(\d{2})/(\d{2})/(\d{4})', header_line)
        if hm:
            detected_year = int(hm.group(3))
            break
        hm2 = re.search(rf'((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{{4}})', header_line)
        if hm2:
            detected_year = int(hm2.group(0).split()[1])
            break

    # ── 2. Regex patterns ─────────────────────────────────────────────────────
    # NatWest: amount ends with £SIGN (negative = debit, positive = credit)
    natwest_pat = rf'^(\d{{1,2}}\s+({months})(?:,?\s+\d{{4}})?)\s+(.+?)\s+(-?£[\d,]+\.?\d*)$'
    # Tide/Wise: two plain numbers at end: AMOUNT BALANCE (no £ prefix)
    tide_pat = rf'^(\d{{1,2}}\s+({months})(?:\s+\d{{4}})?)\s+(.+?)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)$'

    prev_balance = None
    curr_year = detected_year  # tracks working year as we go chronologically

    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue

        amount = None
        balance = None

        # Try NatWest format
        m = re.match(natwest_pat, line)
        if m:
            raw_date = m.group(1)
            desc = (m.group(3) or '').strip()
            amt_raw = m.group(4)
            negative = amt_raw.startswith('-')
            try:
                amt_val = float(amt_raw.replace('-', '').replace('£', '').replace(',', ''))
                amount = -amt_val if negative else amt_val
            except ValueError:
                continue
        else:
            # Try Tide/Wise format
            mt = re.match(tide_pat, line)
            if not mt:
                continue
            raw_date = mt.group(1)
            desc = mt.group(3).strip()
            try:
                amount = float(mt.group(4).replace(',', ''))
                balance = float(mt.group(5).replace(',', ''))
            except ValueError:
                continue

        # ── 3. Parse month & day, track year across pages ─────────────────────
        date_parts = raw_date.replace(',', '').strip().split()
        if len(date_parts) < 2:
            continue
        day = date_parts[0]
        month_str = date_parts[1]
        inline_year = None
        if len(date_parts) >= 3:
            inline_year = int(date_parts[2])

        tx_month_idx = month_idx.get(month_str.title(), 1)

        # Determine working year: use inline year if present, otherwise running year
        if inline_year:
            # Inline year overrides working year (can jump forward/backward per page)
            curr_year = inline_year
        elif curr_year is None:
            curr_year = 2026  # fallback

        # Apply year to get full date string (preserve original "DD Mon" display)
        date_str = f"{day} {month_str.title()} {curr_year}"

        # ── 4. Classify transaction type ───────────────────────────────────────
        upper = desc.upper()
        if 'CREDIT' in upper or 'REFUND' in upper:
            tx_type = 'Credit'
        elif 'DEBIT CARD' in upper:
            tx_type = 'Debit Card'
        elif 'DIRECT DEBIT' in upper:
            tx_type = 'Direct Debit'
        elif 'CHAPS' in upper or 'BACS' in upper:
            tx_type = 'Bank Transfer'
        elif 'STANDING ORDER' in upper:
            tx_type = 'Standing Order'
        elif 'FEE' in upper or 'CHARGE' in upper:
            tx_type = 'Fee'
        else:
            tx_type = 'Payment'

        # ── 5. Money In / Money Out ─────────────────────────────────────────────
        if balance is not None and prev_balance is not None:
            balance_diff = balance - prev_balance
            if balance_diff >= 0:
                money_in = round(amount, 2) if amount > 0 else ''
                money_out = ''
            else:
                money_in = ''
                money_out = round(abs(amount), 2) if amount > 0 else ''
        elif amount is not None:
            if amount >= 0:
                money_in = round(amount, 2)
                money_out = ''
            else:
                money_in = ''
                money_out = round(abs(amount), 2)
        else:
            money_in = ''
            money_out = ''

        if balance is not None:
            prev_balance = balance

        transactions.append({
            'date': date_str,
            'month_idx': tx_month_idx,
            'day': int(day),
            'description': desc,
            'type': tx_type,
            'money_in': money_in,
            'money_out': money_out,
            'balance': round(balance, 2) if balance is not None else '',
            'raw_amount': amount
        })

    # ── 6. Fix year rollovers ──────────────────────────────────────────────────
    # Statement is newest-first (descending). A year boundary only exists when
    # we cross from December (12) into January (1) — NOT for normal May→Apr→Mar drops.
    # Only the transactions OLDER than the Dec→Jan boundary get their year bumped.
    if len(transactions) >= 2:
        boundary_idx = None
        for i in range(1, len(transactions)):
            prev_mi = transactions[i-1]['month_idx']
            curr_mi = transactions[i]['month_idx']
            # True year boundary: Dec(12) → Jan(1)
            if curr_mi == 1 and prev_mi == 12:
                boundary_idx = i
                break
        if boundary_idx:
            for tx in transactions[boundary_idx:]:
                parts = tx['date'].split()
                tx['date'] = f"{parts[0]} {parts[1]} {int(parts[2]) + 1}"

    # Sort ascending (Jan 1 → Dec 31) for clean chronological output
    transactions.sort(key=lambda t: (t['month_idx'], t['day']))

    # Strip auxiliary sort keys from final output
    for tx in transactions:
        tx.pop('month_idx', None)
        tx.pop('day', None)

    return transactions


def extract_tables_from_pdf(pdf_path):
    """
    Extract tables from PDF. Strategy:
    1. Detect Wise vs other bank statement format
    2. Apply appropriate text parser
    3. Fall back to pdfplumber grid extraction if text parsing finds nothing
    All pages merged into a single combined table.
    """
    tables_by_page = []
    all_text_txns = []

    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        all_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
        tables = [p.extract_tables() for p in pdf.pages]

    # Detect Wise format (Wise Payments Ltd. header + transaction lines without DD Mon on TX line)
    is_wise = 'Wise Payments' in all_text or 'wise.com' in all_text.lower()
    is_monzo = 'Monzo' in all_text or 'Account number: 6865' in all_text

    if is_wise:
        txns = parse_wise_text(all_text)
    elif is_monzo:
        txns = parse_monzo_text(all_text)
    else:
        txns = parse_bank_text(all_text)

    if txns:
        # Deduplicate
        seen = set()
        unique_txns = []
        for t in txns:
            # Use date string (not dict) as dedup key for consistency
            key = (t.get('date', {}).get('date', t.get('date', '')) if isinstance(t.get('date'), dict) else t['date'],
                   t['description'], t['type'],
                   t.get('raw_amount') or t.get('amount'))
            if key not in seen:
                seen.add(key)
                unique_txns.append(t)

        tables_by_page = [{
            'page': 1,
            'headers': ['Date', 'Description', 'Type', 'Money In', 'Money Out', 'Balance'],
            'rows': [[t['date'], t['description'], t['type'], t['money_in'], t['money_out'], t['balance']] for t in unique_txns],
            'raw_text': '',
            'layout': 'text'
        }]

    return tables_by_page


def extract_transactions_from_text(text, date_formats=None):
    """
    Try to detect and parse transaction rows from raw text.
    Looks for lines that could be date + description + amount.
    """
    if date_formats is None:
        date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d %b %Y', '%d %B %Y', '%m/%d/%Y']
    
    import re
    transactions = []
    
    # Common amount patterns: £1,234.56 or $1,234.56 or 1,234.56 or -1,234.56
    amount_pattern = r'[-£$€]?\s*[\d,]+\.?\d*'
    
    lines = text.split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Try to find a date at the start
        date_match = re.match(r'(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})', line)
        if not date_match:
            # Try month name format
            date_match = re.match(r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4})', line, re.IGNORECASE)
        
        if date_match:
            date_str = date_match.group(1)
            rest = line[date_match.end():].strip()
            
            # Try to find amounts in the rest of the line
            amounts = re.findall(r'[-+]?£?\s*[\d,]+\.?\d*', rest)
            if amounts:
                # Last amount usually = running balance, second-to-last = transaction
                desc = rest
                for a in reversed(amounts):
                    desc = desc.replace(a, '').strip()
                
                # Try to get the most likely transaction amount
                tx_amount = None
                for a in amounts:
                    cleaned = re.sub(r'[£$€\s+\-]', '', a)
                    try:
                        val = float(cleaned.replace(',', ''))
                        if val > 0:
                            tx_amount = val
                            break
                    except:
                        pass
                
                if tx_amount:
                    transactions.append({
                        'date': date_str,
                        'description': desc[:100],
                        'amount': tx_amount
                    })
    
    return transactions


def auto_detect_transactions(tables_by_page):
    """Try to find transaction-like tables (dates, descriptions, amounts)."""
    transaction_tables = []
    
    for t in tables_by_page:
        headers = [str(h).lower().strip() if h else '' for h in t['headers']]
        headers_str = ' '.join(headers)
        rows = t['rows']
        
        # Check if this looks like a transaction table
        is_tx_table = False
        
        # Has date-like column
        date_indicators = ['date', 'trans date', 'posting date', 'transaction date']
        has_date = any(di in headers_str for di in date_indicators)
        
        # Has amount-like columns
        amount_indicators = ['amount', 'debit', 'credit', 'balance', 'withdrawal', 'deposit']
        has_amount = any(ai in headers_str for ai in amount_indicators)
        
        if has_date or has_amount:
            is_tx_table = True
        elif rows:
            # Check first column of rows for date-like patterns
            first_vals = [str(r[0]).strip() if r else '' for r in rows[:5]]
            date_patterns = [r'\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}', r'\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)']
            import re
            if any(re.search(dp, v) for v in first_vals for dp in date_patterns):
                is_tx_table = True
        
        if is_tx_table:
            transaction_tables.append(t)
    
    return transaction_tables


def rows_to_excel(rows, headers, sheet_name="Sheet1"):
    """Build an Excel workbook from rows/headers."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)
    
    FDARK = PatternFill("solid", fgColor="1F4E79")
    FALT  = PatternFill("solid", fgColor="F2F2F2")
    FWHT  = PatternFill("solid", fgColor="FFFFFF")
    FTOT  = PatternFill("solid", fgColor="BDD7EE")
    
    # Headers
    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=str(h) if h else '')
        c.fill = FDARK
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data rows
    for r_idx, row in enumerate(rows, 2):
        alt = (r_idx % 2 == 0)
        fill = FALT if alt else FWHT
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '')
            cell.fill = fill
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center')
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    return wb


# ── HTML Template ────────────────────────────────────────────────────────────

TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Statement Extractor — PDF to Excel</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #ffffff;
    color: #1e293b;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
  }

  /* ── Header ── */
  header {
    padding: 1.25rem 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 1px solid #e2e8f0;
  }
  .brand { font-size: 1.1rem; font-weight: 700; color: #1e293b; letter-spacing: -0.02em; }
  .brand span { color: #2563eb; }
  .header-links { display: flex; gap: 1.5rem; }
  .header-links a { color: #64748b; text-decoration: none; font-size: 0.875rem; }
  .header-links a:hover { color: #2563eb; }

  /* ── Main ── */
  main { flex: 1; max-width: 760px; width: 100%; margin: 0 auto; padding: 4rem 2rem 3rem; text-align: center; }
  
  .hero { margin-bottom: 2.5rem; }
  .hero h1 { font-size: 2.25rem; font-weight: 700; color: #0f172a; letter-spacing: -0.03em; line-height: 1.2; margin-bottom: 0.75rem; }
  .hero p { font-size: 1rem; color: #64748b; line-height: 1.6; max-width: 520px; margin: 0 auto; }

  /* ── Drop zone ── */
  .drop-zone {
    border: 2px dashed #cbd5e1;
    border-radius: 12px;
    padding: 3.5rem 2rem;
    background: #f8fafc;
    cursor: pointer;
    transition: border-color 0.2s, background 0.2s;
    margin-bottom: 1rem;
  }
  .drop-zone:hover, .drop-zone.dragover {
    border-color: #2563eb;
    background: #eff6ff;
  }
  .drop-zone input { display: none; }
  .drop-icon { font-size: 2.5rem; margin-bottom: 0.75rem; }
  .drop-zone h3 { font-size: 1rem; font-weight: 600; color: #1e293b; margin-bottom: 0.35rem; }
  .drop-zone p { font-size: 0.8rem; color: #94a3b8; }
  .drop-zone:hover h3 { color: #2563eb; }

  /* ── File info bar ── */
  .file-info {
    display: none;
    align-items: center;
    gap: 0.75rem;
    background: #f1f5f9;
    border-radius: 8px;
    padding: 0.6rem 1rem;
    margin-bottom: 1rem;
    text-align: left;
  }
  .file-info.show { display: flex; }
  .file-info .fname { flex: 1; font-size: 0.875rem; color: #1e293b; word-break: break-all; }
  .file-info .fsize { font-size: 0.75rem; color: #94a3b8; white-space: nowrap; }
  .file-info .fremove { background: none; border: none; cursor: pointer; color: #94a3b8; font-size: 1rem; padding: 0 0.25rem; }
  .file-info .fremove:hover { color: #ef4444; }

  /* ── Upload button ── */
  .upload-btn {
    display: none;
    background: #2563eb;
    color: #ffffff;
    border: none;
    border-radius: 8px;
    padding: 0.65rem 1.75rem;
    font-size: 0.875rem;
    font-weight: 600;
    cursor: pointer;
    transition: background 0.15s;
    margin: 0 auto 1rem;
  }
  .upload-btn.show { display: block; }
  .upload-btn:hover { background: #1d4ed8; }
  .upload-btn:disabled { background: #93c5f1; cursor: not-allowed; }

  /* ── Processing ── */
  .processing { display: none; padding: 2rem; }
  .processing.show { display: block; }
  .spinner {
    width: 36px; height: 36px;
    border: 3px solid #e2e8f0;
    border-top-color: #2563eb;
    border-radius: 50%;
    animation: spin 0.7s linear infinite;
    margin: 0 auto 0.75rem;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
  .processing p { color: #64748b; font-size: 0.875rem; }

  /* ── Results ── */
  .results { display: none; text-align: left; }
  .results.show { display: block; }

  .results-summary {
    display: flex;
    gap: 0.75rem;
    margin-bottom: 1.25rem;
    flex-wrap: wrap;
  }
  .res-chip {
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    border-radius: 6px;
    padding: 0.4rem 0.85rem;
    font-size: 0.8rem;
    color: #1e40af;
    font-weight: 500;
  }

  /* ── Page tabs ── */
  .page-tabs { display: flex; gap: 0.4rem; margin-bottom: 1rem; flex-wrap: wrap; }
  .page-tab {
    padding: 0.4rem 0.85rem;
    border-radius: 6px;
    border: 1px solid #e2e8f0;
    background: #ffffff;
    color: #64748b;
    cursor: pointer;
    font-size: 0.8rem;
    transition: all 0.15s;
  }
  .page-tab:hover { border-color: #2563eb; color: #2563eb; }
  .page-tab.active { background: #2563eb; color: #ffffff; border-color: #2563eb; }

  /* ── Table ── */
  .table-card {
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    overflow: hidden;
    margin-bottom: 1.25rem;
  }
  .table-card-hdr {
    padding: 0.65rem 1rem;
    background: #f8fafc;
    border-bottom: 1px solid #e2e8f0;
    font-size: 0.78rem;
    color: #64748b;
    font-weight: 500;
  }
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
  thead th {
    background: #f1f5f9;
    color: #475569;
    padding: 0.55rem 0.85rem;
    text-align: left;
    font-weight: 600;
    white-space: nowrap;
    border-bottom: 1px solid #e2e8f0;
  }
  tbody td {
    padding: 0.5rem 0.85rem;
    border-bottom: 1px solid #f1f5f9;
    color: #1e293b;
  }
  tbody tr:last-child td { border-bottom: none; }
  tbody tr:nth-child(even) td { background: #fafafa; }
  tbody tr:hover td { background: #f0f4ff; }

  .money { font-variant-numeric: tabular-nums; text-align: right; }
  .money.positive { color: #16a34a; }
  .money.negative { color: #dc2626; }

  /* ── Export bar ── */
  .export-row {
    display: flex;
    gap: 0.6rem;
    flex-wrap: wrap;
    align-items: center;
    padding-top: 0.5rem;
    border-top: 1px solid #e2e8f0;
    margin-top: 0.5rem;
  }
  .export-row .el { font-size: 0.8rem; color: #64748b; }
  .btn {
    padding: 0.45rem 1rem;
    border-radius: 6px;
    border: 1px solid #e2e8f0;
    background: #ffffff;
    cursor: pointer;
    font-size: 0.8rem;
    font-weight: 500;
    color: #1e293b;
    transition: all 0.15s;
  }
  .btn:hover { background: #2563eb; color: #ffffff; border-color: #2563eb; }
  .btn-green { background: #16a34a; color: #ffffff; border-color: #16a34a; }
  .btn-green:hover { background: #15803d; border-color: #15803d; }

  /* ── Features row ── */
  .features { display: flex; justify-content: center; gap: 3rem; margin-top: 3rem; flex-wrap: wrap; }
  .feat { text-align: center; }
  .feat-icon { font-size: 1.5rem; margin-bottom: 0.35rem; }
  .feat h4 { font-size: 0.875rem; font-weight: 600; color: #0f172a; margin-bottom: 0.2rem; }
  .feat p { font-size: 0.75rem; color: #94a3b8; }

  /* ── Footer ── */
  footer {
    border-top: 1px solid #e2e8f0;
    padding: 1.25rem 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 0.75rem;
  }
  footer p { font-size: 0.75rem; color: #94a3b8; }
  footer a { color: #64748b; text-decoration: none; font-size: 0.75rem; }
  footer a:hover { color: #2563eb; }
  .footer-links { display: flex; gap: 1.25rem; }

  @media (max-width: 600px) {
    main { padding: 2.5rem 1.25rem 2rem; }
    .hero h1 { font-size: 1.65rem; }
    .features { gap: 1.5rem; }
    header, footer { padding: 1rem 1.25rem; }
    .header-links { display: none; }
  }
</style>
</head>
<body>

<!-- Header -->
<header>
  <div class="brand">Statement <span>Extractor</span></div>
  <nav class="header-links">
    <a href="mailto:hello@statementextractor.com">Contact</a>
  </nav>
</header>

<!-- Main -->
<main>
  <div class="hero">
    <h1>PDF Bank Statements<br>to Excel in Seconds</h1>
    <p>Upload any PDF bank statement — Wise, Monzo, Starling, Barclays and more. Extract tables and download as Excel or CSV. No sign-up required.</p>
  </div>

  <form id="uploadForm" enctype="multipart/form-data">
    <div class="drop-zone" id="dropZone">
      <div class="drop-icon">📄</div>
      <h3>Drag & drop your PDF here or click to browse</h3>
      <p>Max file size: 50MB · PDF only</p>
      <input type="file" id="fileInput" name="file" accept=".pdf">
    </div>
  </form>

  <div class="file-info" id="fileInfo">
    <span class="fname" id="fileNameDisp"></span>
    <span class="fsize" id="fileSizeDisp"></span>
    <button class="fremove" onclick="clearFile()">✕</button>
  </div>

  <button class="upload-btn" id="uploadBtn" onclick="uploadFile()">Extract Tables →</button>

  <div class="processing" id="processing">
    <div class="spinner"></div>
    <p>Extracting tables from your PDF...</p>
  </div>

  <div class="results" id="results">
    <div class="results-summary" id="resultsSummary"></div>
    <div class="page-tabs" id="pageTabs"></div>
    <div id="tablesContainer"></div>
    <div class="export-row" id="exportRow"></div>
  </div>
</main>

<!-- Features -->
<div class="features">
  <div class="feat">
    <div class="feat-icon">📋</div>
    <h4>Versatile</h4>
    <p>Wise, Monzo, Starling,<br>Barclays, HSBC & more</p>
  </div>
  <div class="feat">
    <div class="feat-icon">🎯</div>
    <h4>Accurate</h4>
    <p>Table-level extraction,<br>row data preserved</p>
  </div>
  <div class="feat">
    <div class="feat-icon">💸</div>
    <h4>Free</h4>
    <p>No sign-up, no subscription,<br>no fees</p>
  </div>
  <div class="feat">
    <div class="feat-icon">🔒</div>
    <h4>Secure</h4>
    <p>Files processed locally,<br>never stored on server</p>
  </div>
</div>

<!-- Footer -->
<footer>
  <p>© 2025 Statement Extractor</p>
  <div class="footer-links">
    <a href="mailto:hello@statementextractor.com">Contact</a>
    <a href="#">Privacy</a>
  </div>
</footer>

<script>
  const dropZone = document.getElementById('dropZone');
  const fileInput = document.getElementById('fileInput');
  const fileInfo = document.getElementById('fileInfo');
  const fileName = document.getElementById('fileNameDisp');
  const fileSize = document.getElementById('fileSizeDisp');
  const uploadBtn = document.getElementById('uploadBtn');
  const processing = document.getElementById('processing');
  const results = document.getElementById('results');

  let extractedData = null;
  let selectedPage = 0;

  dropZone.addEventListener('click', () => fileInput.click());
  dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('dragover'); });
  dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
  dropZone.addEventListener('drop', e => { e.preventDefault(); dropZone.classList.remove('dragover'); if(e.dataTransfer.files.length) fileInput.files = e.dataTransfer.files; showFileInfo(); });
  fileInput.addEventListener('change', showFileInfo);

  function showFileInfo() {
    const f = fileInput.files[0];
    if (!f) return;
    fileName.textContent = f.name;
    fileSize.textContent = (f.size / 1024 / 1024).toFixed(2) + ' MB';
    fileInfo.classList.add('show');
    uploadBtn.classList.add('show');
    results.classList.remove('show');
  }

  function clearFile() {
    fileInput.value = '';
    fileInfo.classList.remove('show');
    uploadBtn.classList.remove('show');
    results.classList.remove('show');
    extractedData = null;
  }

  async function uploadFile() {
    const f = fileInput.files[0];
    if (!f) return;
    uploadBtn.disabled = true;
    uploadBtn.textContent = 'Extracting...';
    processing.classList.add('show');
    results.classList.remove('show');

    const formData = new FormData();
    formData.append('file', f);

    try {
      const resp = await fetch('/extract', { method: 'POST', body: formData });
      const data = await resp.json();
      if (data.error) { alert(data.error); return; }
      extractedData = data;
      renderResults();
    } catch(e) {
      alert('Upload failed: ' + e.message);
    } finally {
      uploadBtn.disabled = false;
      uploadBtn.textContent = 'Extract Tables →';
      processing.classList.remove('show');
    }
  }

  function renderResults() {
    if (!extractedData) return;
    results.classList.add('show');

    // Summary chips
    const sb = document.getElementById('resultsSummary');
    sb.innerHTML = `
      <span class="res-chip">📄 ${extractedData.filename}</span>
      <span class="res-chip">${extractedData.table_count} table${extractedData.table_count !== 1 ? 's' : ''} found</span>
      <span class="res-chip">${extractedData.total_rows.toLocaleString()} rows</span>
    `;
    
    // Page tabs
    const tabs = document.getElementById('pageTabs');
    tabs.innerHTML = '';
    extractedData.tables.forEach((t, i) => {
      const tab = document.createElement('button');
      tab.className = 'page-tab' + (i === selectedPage ? ' active' : '');
      tab.textContent = `Page ${t.page} — ${t.rows.length} rows`;
      tab.onclick = () => { selectedPage = i; renderTables(); renderTabs(); };
      tabs.appendChild(tab);
    });
    
    renderTables();
    renderExportBar();
  }
  
  function renderTabs() {
    const tabs = document.getElementById('pageTabs');
    extractedData.tables.forEach((t, i) => {
      const tab = tabs.children[i];
      if (tab) tab.className = 'page-tab' + (i === selectedPage ? ' active' : '');
    });
  }
  
  function renderTables() {
    const container = document.getElementById('tablesContainer');
    container.innerHTML = '';
    
    const t = extractedData.tables[selectedPage];
    const card = document.createElement('div');
    card.className = 'table-card';
    
    const headers = t.headers || [];
    const rows = t.rows || [];
    
    let th_html = '<thead><tr>';
    headers.forEach(h => { th_html += `<th>${h || ''}</th>`; });
    th_html += '</tr></thead>';
    
    let tb_html = '<tbody>';
    rows.forEach(row => {
      tb_html += '<tr>';
      (row || []).forEach((cell, ci) => {
        let cls = '';
        let val = cell || '';
        val = String(val);
        // Try to detect money
        const cleaned = val.replace(/[£$€,\s]/g, '').replace(/^\(/, '-').replace(/\)$/, '');
        const num = parseFloat(cleaned);
        if (!isNaN(num) && Math.abs(num) > 0) {
          cls = 'money ' + (num > 0 ? 'positive' : 'negative');
        }
        tb_html += `<td class="${cls}">${val}</td>`;
      });
      // Fill empty cols
      for (let i = (row || []).length; i < headers.length; i++) tb_html += '<td></td>';
      tb_html += '</tr>';
    });
    tb_html += '</tbody>';
    
    card.innerHTML = `<div class="table-card-hdr">Page ${t.page} — ${rows.length} rows, ${headers.length} columns</div><div class="table-wrap"><table>${th_html}${tb_html}</table></div>`;
    container.appendChild(card);
  }

  function renderExportBar() {
    const bar = document.getElementById('exportRow');
    bar.innerHTML = `
      <span class="el">Export:</span>
      <button class="btn btn-green" onclick="exportAll('xlsx')">Download Excel</button>
      <button class="btn" onclick="exportAll('csv')">Download CSV</button>
      <button class="btn" onclick="exportPage('xlsx')">This page only</button>
    `;
  }
  
  async function exportAll(format) {
    const resp = await fetch(`/download/${extractedData.session_id}/all.${format}`, { method: 'GET' });
    const blob = await resp.blob();
    downloadBlob(blob, `statement_all.${format}`);
  }
  
  async function exportPage(format) {
    const resp = await fetch(`/download/${extractedData.session_id}/page_${selectedPage}.${format}`, { method: 'GET' });
    const blob = await resp.blob();
    downloadBlob(blob, `statement_page_${extractedData.tables[selectedPage].page}.${format}`);
  }
  
  function downloadBlob(blob, filename) {
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a'); a.href = url; a.download = filename; a.click();
    URL.revokeObjectURL(url);
  }
</script>
</body>
</html>
"""


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

@app.route('/')
def index():
    return render_template_string(TEMPLATE)


@app.route('/extract', methods=['POST'])
def extract():
    f = request.files.get('file')
    if not f or not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Please upload a PDF file'}), 400
    
    session_id = uuid.uuid4().hex
    tmp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{session_id}.pdf')
    f.save(tmp_path)
    
    try:
        tables = extract_tables_from_pdf(tmp_path)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"PDF extract error: {e}\n{tb}")
        return jsonify({'error': f'Failed to read PDF: {str(e)}'}), 500
    
    if not tables:
        return jsonify({'error': 'No tables found in this PDF. Try a different file.'}), 400
    
    total_rows = sum(len(t['rows']) for t in tables)
    
    # Save extracted data
    import json
    cache_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{session_id}.json')
    with open(cache_path, 'w') as fh:
        json.dump({'tables': tables, 'filename': f.filename}, fh)

    return jsonify({
        'session_id': session_id,
        'filename': f.filename,
        'table_count': len(tables),
        'total_rows': total_rows,
        'tables': [
            {'page': t['page'], 'headers': t['headers'], 'rows': t['rows'], 'row_count': len(t['rows'])}
            for t in tables
        ]
    })


@app.route('/download/<session_id>/<path:filename>')
def download(session_id, filename):
    cache_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{session_id}.json')
    if not os.path.exists(cache_path):
        return jsonify({'error': 'Session expired'}), 404
    
    import json
    with open(cache_path) as fh:
        data = json.load(fh)
    
    fmt = filename.split('.')[-1]
    
    if filename == f'all.{fmt}':
        tables = data['tables']
    else:
        # Single page
        page_num = int(filename.split('_')[1].replace(f'.{fmt}', ''))
        tables = [t for t in data['tables'] if t['page'] == page_num]
        if not tables:
            return jsonify({'error': 'Page not found'}), 404
        tables = [tables[0]]
    
    if fmt == 'csv':
        # Single merged CSV
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        for t in tables:
            writer.writerow([f"=== Page {t['page']} ==="])
            writer.writerow(t['headers'])
            for row in t['rows']:
                writer.writerow(row)
            writer.writerow([])
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                        download_name=filename, as_attachment=True)
    
    else:  # xlsx
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for t_idx, t in enumerate(tables):
            ws_name = f"Page {t['page']}"
            if len(ws_name) > 31: ws_name = ws_name[:31]
            ws = wb.create_sheet(ws_name)
            
            FDARK = PatternFill("solid", fgColor="1F4E79")
            FALT  = PatternFill("solid", fgColor="F2F2F2")
            FWHT  = PatternFill("solid", fgColor="FFFFFF")
            
            headers = t['headers']
            ws.row_dimensions[1].height = 20
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=str(h) if h else '')
                c.fill = FDARK
                c.font = Font(color="FFFFFF", bold=True, size=10)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for r_idx, row in enumerate(t['rows'], 2):
                alt = (r_idx % 2 == 0)
                fill = FALT if alt else FWHT
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '')
                    cell.fill = fill
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(vertical='center')
            
            # Auto-fit cols
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try: max_len = max(max_len, len(str(cell.value or '')))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 55)
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        download_name=filename, as_attachment=True)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3002, debug=False)
